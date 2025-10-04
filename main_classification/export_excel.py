import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

def export_results_to_excel(results: dict, output_path: str) -> bool:
    """
    Xuất kết quả phân loại ra file Excel
    
    Args:
        results: Dictionary chứa kết quả phân loại
        output_path: Đường dẫn file Excel đầu ra
    
    Returns:
        bool: True nếu thành công, False nếu thất bại
    """
    try:
        # Tạo workbook mới
        wb = Workbook()
        
        # Xóa sheet mặc định
        wb.remove(wb.active)
        
        # Định nghĩa màu sắc cho từng loại
        colors = {
            "hoạt động": "00FF00",      # Xanh lá
            "leave_message": "FFFF00",   # Vàng
            "be_blocked": "FF0000",     # Đỏ
            "can_not_connect": "FFA500", # Cam
            "incorrect": "800080",      # Tím
            "ringback_tone": "00FFFF",   # Cyan
            "waiting_tone": "FFC0CB",   # Hồng
            "mute": "808080"            # Xám
        }
        
        # Tạo sheet tổng hợp
        summary_sheet = wb.create_sheet("Tổng Hợp")
        _create_summary_sheet(summary_sheet, results, colors)
        
        # Tạo sheet chi tiết cho từng loại
        for category, data_list in results.items():
            if data_list:  # Chỉ tạo sheet nếu có dữ liệu
                sheet = wb.create_sheet(category.title())
                _create_detail_sheet(sheet, category, data_list, colors.get(category, "FFFFFF"))
        
        # Lưu file
        wb.save(output_path)
        logger.info(f"✅ Đã xuất kết quả ra file: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi xuất Excel: {e}")
        return False

def _create_summary_sheet(sheet, results: dict, colors: dict):
    """Tạo sheet tổng hợp"""
    
    # Tiêu đề
    sheet['A1'] = "BÁO CÁO TỔNG HỢP PHÂN LOẠI SỐ ĐIỆN THOẠI"
    sheet['A1'].font = Font(size=16, bold=True, color="000000")
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:C1')
    
    # Thời gian tạo báo cáo
    sheet['A2'] = f"Thời gian tạo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A2'].font = Font(size=10, italic=True)
    sheet.merge_cells('A2:C2')
    
    # Khoảng trống
    sheet['A4'] = ""
    
    # Bảng thống kê
    sheet['A5'] = "STT"
    sheet['B5'] = "Loại"
    sheet['C5'] = "Số lượng"
    sheet['D5'] = "Tỷ lệ (%)"
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col in ['A5', 'B5', 'C5', 'D5']:
        sheet[col].font = header_font
        sheet[col].fill = header_fill
        sheet[col].alignment = header_alignment
    
    # Tính tổng số lượng
    total_count = sum(len(data_list) for data_list in results.values())
    
    # Dữ liệu thống kê
    row = 6
    for i, (category, data_list) in enumerate(results.items(), 1):
        count = len(data_list)
        percentage = (count / total_count * 100) if total_count > 0 else 0
        
        sheet[f'A{row}'] = i
        sheet[f'B{row}'] = category.replace('_', ' ').title()
        sheet[f'C{row}'] = count
        sheet[f'D{row}'] = f"{percentage:.1f}%"
        
        # Màu nền cho từng loại
        if category in colors:
            fill_color = colors[category]
            sheet[f'B{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Alignment
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')
        sheet[f'C{row}'].alignment = Alignment(horizontal='center')
        sheet[f'D{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Tổng cộng
    sheet[f'A{row}'] = "TỔNG CỘNG"
    sheet[f'B{row}'] = ""
    sheet[f'C{row}'] = total_count
    sheet[f'D{row}'] = "100.0%"
    
    # Style cho tổng cộng
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    for col in ['A', 'C', 'D']:
        sheet[f'{col}{row}'].font = total_font
        sheet[f'{col}{row}'].fill = total_fill
        sheet[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Điều chỉnh độ rộng cột
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12

def _create_detail_sheet(sheet, category: str, data_list: list, color: str):
    """Tạo sheet chi tiết cho một loại"""
    
    # Tiêu đề
    sheet['A1'] = f"CHI TIẾT - {category.replace('_', ' ').upper()}"
    sheet['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A1:H1')
    
    # Header bảng
    headers = ["STT", "Cổng GSM", "Số Điện Thoại", "Thời Gian", "File Audio", "Nội Dung", "Lỗi", "Ghi Chú"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dữ liệu
    for i, data in enumerate(data_list, 1):
        row = i + 3
        
        sheet[f'A{row}'] = i
        sheet[f'B{row}'] = data.get('port', 'N/A')
        sheet[f'C{row}'] = data.get('phone', 'N/A')
        sheet[f'D{row}'] = data.get('timestamp', 'N/A')
        sheet[f'E{row}'] = data.get('file', 'N/A')
        sheet[f'F{row}'] = data.get('transcription', 'N/A')
        sheet[f'G{row}'] = data.get('error', 'N/A')
        
        # Ghi chú đặc biệt
        note = ""
        if data.get('status') == 'hoạt động':
            note = "Người nghe đã nhấc máy"
        elif data.get('error'):
            note = "Có lỗi xảy ra"
        elif not data.get('transcription'):
            note = "Không có nội dung audio"
        
        sheet[f'H{row}'] = note
        
        # Màu xen kẽ cho các hàng
        if i % 2 == 0:
            for col in range(1, 9):
                sheet.cell(row=row, column=col).fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )
    
    # Điều chỉnh độ rộng cột
    column_widths = {
        'A': 8,   # STT
        'B': 12,  # Cổng GSM
        'C': 15,  # Số điện thoại
        'D': 20,  # Thời gian
        'E': 25,  # File Audio
        'F': 40,  # Nội dung
        'G': 20,  # Lỗi
        'H': 25   # Ghi chú
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # Wrap text cho cột nội dung và ghi chú
    for row in range(4, len(data_list) + 4):
        sheet[f'F{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        sheet[f'H{row}'].alignment = Alignment(wrap_text=True, vertical='top')

def create_sample_results():
    """Tạo dữ liệu mẫu để test"""
    return {
        "hoạt động": [
            {
                "port": "COM37",
                "phone": "0987654321",
                "timestamp": "20241201_143022",
                "file": None,
                "transcription": "Người nghe đã nhấc máy",
                "error": None
            }
        ],
        "leave_message": [
            {
                "port": "COM38",
                "phone": "0987654322",
                "timestamp": "20241201_143025",
                "file": "0987654322_20241201_143025.amr",
                "transcription": "quý khách vui lòng để lại lời nhắn sau tiếng bíp",
                "error": None
            }
        ],
        "be_blocked": [
            {
                "port": "COM39",
                "phone": "0987654323",
                "timestamp": "20241201_143028",
                "file": "0987654323_20241201_143028.amr",
                "transcription": "đang tạm khóa khuyên quý khách vui lòng tải lại",
                "error": None
            }
        ],
        "can_not_connect": [
            {
                "port": "COM40",
                "phone": "0987654324",
                "timestamp": "20241201_143031",
                "file": None,
                "transcription": None,
                "error": "Không thể gọi điện"
            }
        ]
    }

if __name__ == "__main__":
    # Test với dữ liệu mẫu
    sample_results = create_sample_results()
    success = export_results_to_excel(sample_results, "test_results.xlsx")
    
    if success:
        print("✅ Test xuất Excel thành công!")
    else:
        print("❌ Test xuất Excel thất bại!")
