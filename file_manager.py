import os
from typing import List, Tuple
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class FileManager:
    def __init__(self):
        self.phone_numbers = []
        # 3 kết quả chung cho tất cả các số
        self.hoat_dong = []      # HOẠT ĐỘNG
        self.thue_bao = []       # THUÊ BAO  
        self.so_khong_dung = []  # SỐ KHÔNG ĐÚNG
        
    def load_phone_numbers(self, file_path: str) -> List[str]:
        """Đọc danh sách số điện thoại từ file txt"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                numbers = [line.strip() for line in file if line.strip()]
            self.phone_numbers = numbers
            return numbers
        except Exception as e:
            raise Exception(f"Lỗi đọc file: {str(e)}")
    
    def save_results(self, output_dir: str = "results"):
        """Lưu kết quả vào file Excel với 3 cột chung: HOẠT ĐỘNG, THUÊ BAO, SỐ KHÔNG ĐÚNG"""
        # Tạo thư mục results nếu chưa có
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Tạo file Excel
        excel_file = os.path.join(output_dir, f"gsm_check_results_{timestamp}.xlsx")
        
        # Tạo workbook và worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Kết quả check số điện thoại"
        
        # Định dạng header
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Tạo header cho 3 kết quả chung
        ws['A1'] = "HOẠT ĐỘNG"
        ws['B1'] = "THUÊ BAO"
        ws['C1'] = "SỐ KHÔNG ĐÚNG"
        
        # Áp dụng định dạng cho header
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = header_alignment
        
        # Điền dữ liệu cho 3 kết quả chung
        max_rows = max(len(self.hoat_dong), len(self.thue_bao), len(self.so_khong_dung))
        
        for i in range(max_rows):
            row = i + 2  # Bắt đầu từ row 2 (sau header)
            
            # Cột HOẠT ĐỘNG
            if i < len(self.hoat_dong):
                ws[f'A{row}'] = self.hoat_dong[i]
            
            # Cột THUÊ BAO
            if i < len(self.thue_bao):
                ws[f'B{row}'] = self.thue_bao[i]
            
            # Cột SỐ KHÔNG ĐÚNG
            if i < len(self.so_khong_dung):
                ws[f'C{row}'] = self.so_khong_dung[i]
        
        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # Lưu file
        wb.save(excel_file)
        
        return excel_file
    
    def add_result(self, phone_number: str, result_type: str):
        """Thêm kết quả check cho một số điện thoại với 3 loại: HOẠT ĐỘNG, THUÊ BAO, SỐ KHÔNG ĐÚNG"""
        if result_type == "HOẠT ĐỘNG":
            self.hoat_dong.append(phone_number)
        elif result_type == "THUÊ BAO":
            self.thue_bao.append(phone_number)
        elif result_type == "SỐ KHÔNG ĐÚNG":
            self.so_khong_dung.append(phone_number)
    
    def get_progress(self) -> Tuple[int, int, int]:
        """Trả về tiến độ: đã check, tổng số, phần trăm"""
        total = len(self.phone_numbers)
        checked = len(self.hoat_dong) + len(self.thue_bao) + len(self.so_khong_dung)
        percentage = (checked / total * 100) if total > 0 else 0
        return checked, total, percentage
    
    def reset_results(self):
        """Reset kết quả để bắt đầu check mới"""
        self.hoat_dong = []
        self.thue_bao = []
        self.so_khong_dung = [] 